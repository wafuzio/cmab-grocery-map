import openpyxl
import json
import re

wb = openpyxl.load_workbook('/Users/dan.maguire/Downloads/WMT Stores/2026 MilkPEP Neptune Supergirl Promo.xlsx')

stores = []

# Extract Safeway/Albertsons stores from Crystal Creamery - Safeway
print("Processing Crystal Creamery - Safeway...")
ws = wb['Crystal Creamery - Safeway']
for row in range(3, ws.max_row + 1):
    name = ws.cell(row, 2).value
    addr = ws.cell(row, 3).value
    city = ws.cell(row, 4).value
    state = ws.cell(row, 5).value
    zip_code = ws.cell(row, 6).value
    
    if name and city and state:
        # Extract store number from name
        store_num = None
        match = re.search(r'#(\d+)', str(name))
        if match:
            store_num = match.group(1)
        
        stores.append({
            'retailer': 'Albertsons',
            'processor': 'Crystal Creamery',
            'store_num': store_num or 'UNK',
            'address': str(addr) if addr else '',
            'city': str(city),
            'state': str(state),
            'zip': str(zip_code).replace('-0000', '') if zip_code else ''
        })

print(f"Found {len(stores)} Crystal Creamery - Safeway stores")

# Extract Lucerne stores (Albertsons brand)
print("\nProcessing Lucerne...")
ws = wb['Lucerne']
lucerne_count = 0
for row in range(2, ws.max_row + 1):
    banner = ws.cell(row, 3).value
    facility = ws.cell(row, 4).value
    addr = ws.cell(row, 7).value
    city = ws.cell(row, 8).value
    state = ws.cell(row, 9).value
    zip_code = ws.cell(row, 10).value
    
    if city and state and addr:
        stores.append({
            'retailer': 'Albertsons',
            'processor': 'Lucerne',
            'store_num': str(int(facility)) if facility else 'UNK',
            'address': str(addr),
            'city': str(city),
            'state': str(state),
            'zip': str(zip_code) if zip_code else ''
        })
        lucerne_count += 1

print(f"Found {lucerne_count} Lucerne stores")

# Extract Hollandia - Albertsons
print("\nProcessing Hollandia - Albertsons...")
ws = wb['Hollandia - Albertsons']
hollandia_count = 0
for row in range(3, ws.max_row + 1):
    name = ws.cell(row, 1).value
    addr = ws.cell(row, 2).value
    city = ws.cell(row, 3).value
    state = ws.cell(row, 4).value
    zip_code = ws.cell(row, 5).value
    
    if name and city and state:
        # Extract store number
        store_num = None
        match = re.search(r'#(\d+)', str(name))
        if match:
            store_num = match.group(1)
        
        stores.append({
            'retailer': 'Albertsons',
            'processor': 'Hollandia',
            'store_num': store_num or 'UNK',
            'address': str(addr) if addr else '',
            'city': str(city),
            'state': str(state),
            'zip': str(int(zip_code)) if zip_code else ''
        })
        hollandia_count += 1

print(f"Found {hollandia_count} Hollandia - Albertsons stores")

print(f"\nTotal stores to add: {len(stores)}")
print(f"\nSample stores:")
for i in range(min(5, len(stores))):
    print(f"  {stores[i]}")

# Save to JSON for review
with open('/Users/dan.maguire/Downloads/WMT Stores/new_retailers.json', 'w') as f:
    json.dump(stores, indent=2, fp=f)

print(f"\nSaved to new_retailers.json")
